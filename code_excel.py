import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
import time
import tkinter 
import pandas as pd
from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.workbook.protection import WorkbookProtection
global open_path, save_path, save, CSDL_path, BM_path
global wsCSDLKL, wsDonGia, wsBG, wsDM_T_C, BG_max_row_1
global wbkl, wbcsdl, wb_BG, wskl,wscb,wsBMS, wsVTTC
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
# Hệ số lời
HS_VT = 1.0
HS_NC = 1.0
def got_data_from(open_path, CSDL_path, BM_path):
	global wbkl, wbcsdl, wb_BG, wskl,wscb,wsBMS, wsVTTC
	path = open_path
	csdl = CSDL_path
	BG_form = BM_path
	# Kích hoạt workbook
	#wbkl = openpyxl.load_workbook(path)

	wbkl = openpyxl.load_workbook(path)
	wbcsdl = openpyxl.load_workbook(csdl)
	wb_BG = openpyxl.load_workbook(BG_form)

	# Kích hoạt worksheet
	wskl = wbkl["PLBMS"]
	wscb = wbkl["CBBMS"]
	wsBMS = wbkl["WSBMS"]
	wsVTTC = wbkl["VTTC"]
	global wsCSDLKL, wsDonGia, wsBG, wsDM_T_C, BG_max_row_1
	wsCSDLKL = wbcsdl["PLKL"]
	wsDonGia = wbcsdl["CSDLBMS"]
	wsBG = wb_BG["BMS"]
	wsDM_T_C = wbcsdl["DMTC"]
	# Tìm max row trong sheet
	BG_max_row_1 = wsBG.max_row
	# Tìm model trong cơ sở dữ liệu 
def wordfinder(searchString, Soluong, BG_max_row):
	global wsCSDLKL, wsDonGia, wsBG, wsDM_T_C, BG_max_row_1
	global wbkl, wbcsdl, wb_BG, wskl,wscb,wsBMS, wsVTTC
	for i in range(1, wsDonGia.max_row + 1):
		for j in range(1, wsDonGia.max_column + 1):
			if searchString == wsDonGia.cell(i,j).value:
				wsBG.cell(BG_max_row, 2).value = wsDonGia.cell(i,j-1).value
				wsBG.cell(BG_max_row, 3).value = wsDonGia.cell(i,j+1).value
				wsBG.cell(BG_max_row, 4).value = wsDonGia.cell(i,j+2).value
				wsBG.cell(BG_max_row, 5).value = Soluong
				#Code vật tư
				wsBG.cell(BG_max_row, 12).value = wsDonGia.cell(i,j-2).value
				# Đơn giá vật tư
				wsBG.cell(BG_max_row, 14).value = wsDonGia.cell(i,j+3).value
				# Đơn giá nhân công
				wsBG.cell(BG_max_row, 15).value = wsDonGia.cell(i,j+4).value
				# Hệ số vật tư
				wsBG.cell(BG_max_row, 16).value = HS_VT
				# Hệ số nhân công
				wsBG.cell(BG_max_row, 17).value = HS_NC
				# Chèn công thức vật tư
				str1 = wsBG.cell(BG_max_row, 14)
				str2 = str(str1)
				str3 = "="
				str4 = "*"
				str8 = wsBG.cell(BG_max_row, 16)
				str6 = str(str8)
				str5 = str3 + str2[12:-1] + str4 + str6[12:-1]
				wsBG.cell(BG_max_row, 18).value = str5
				# Chèn công thức nhân công
				str11 = wsBG.cell(BG_max_row, 15)
				str12 = str(str11)
				str13 = "="
				str14 = "*"
				str18 = wsBG.cell(BG_max_row, 17)
				str16 = str(str18)
				str15 = str13 + str12[12:-1] + str14 + str16[12:-1]
				wsBG.cell(BG_max_row, 19).value = str15
				# Nhập công thức cột đơn giá vật tư
				str21 = str(wsBG.cell(BG_max_row, 18))
				str22 = str21[12:-1]
				str23 = "="
				wsBG.cell(BG_max_row, 6).value = str23 + str22
				# Nhập công thức cột đơn giá nhân công
				str31 = str(wsBG.cell(BG_max_row, 19))
				str32 = str31[12:-1]
				str33 = "="
				wsBG.cell(BG_max_row, 7).value = str33 + str32
				# Nhập công thức cột thành tiền vật tư
				str41 = "="
				str42 = "*"
				str43 = str(wsBG.cell(BG_max_row, 5))
				str44 = str(wsBG.cell(BG_max_row, 6))
				str45 = str41 + str43[12:-1] + str42 + str44[12:-1]
				wsBG.cell(BG_max_row, 8).value = str45
				#print(str45)

				# Nhập công thức cột thành tiền nhân công
				str51 = "="
				str52 = "*"
				str53 = str(wsBG.cell(BG_max_row, 5))
				str54 = str(wsBG.cell(BG_max_row, 7))
				str55 = str51 + str53[12:-1] + str52 + str54[12:-1]
				wsBG.cell(BG_max_row, 9).value = str55

				# Nhập công thức cột thành tiền tổng cộng
				str61 = "="
				str62 = "+"
				str63 = str(wsBG.cell(BG_max_row, 8))
				str64 = str(wsBG.cell(BG_max_row, 9))
				str65 = str61 + str63[12:-1] + str62 + str64[12:-1]
				wsBG.cell(BG_max_row, 10).value = str65
				BG_max_row = wsBG.max_row + 1
def main_code(save_path, DA_name):
	#Bộ điều khiển
	#Cảm biến
	global wsCSDLKL, wsDonGia, wsBG, wsDM_T_C, BG_max_row_1
	global wbkl, wbcsdl, wb_BG, wskl,wscb,wsBMS, wsVTTC
	global open_path, save
	wsBG.cell(BG_max_row_1, 2).value = "Tủ điều khiển"
	wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
	BG_max_row_1 = wsBG.max_row + 1
	for i in range(15, wskl.max_row+1):
		Pdi = wskl.cell(i,3).value
		Pdo = wskl.cell(i,4).value
		Pai = wskl.cell(i,5).value
		Pao = wskl.cell(i,6).value
		Pui = Pdi + Pai
		Puo = Pdo + Pao
		if Puo <= 8 and Pui <= 12:
			MPS = 1
			MPPIO = 0
			MPPI = 0
		elif Puo <= 8 and Pui > 12 and Pui <= 36:
			MPS = 1
			MPPIO = 0
			MPPI = 1
		elif Puo <= 8 and Pui > 36 and Pui <= 60:
			MPS = 1
			MPPIO = 0
			MPPI = 2
		elif Puo > 8 and Puo <= 16 and Pui >= 12 and Pui <= 24:
			MPS = 1
			MPPIO = 1
			MPPI = 0
		elif Puo > 8 and Puo <= 16 and Pui > 24 and Pui <= 48:
			MPS = 1
			MPPIO = 1 
			MPPI = 1
		elif Puo > 8 and Puo <= 16 and Pui > 48 and Pui <= 72:
			MPS = 1
			MPPIO = 1 
			MPPI = 2
		elif Puo > 16 and Puo <= 24 and Pui >= 12 and Pui <= 36:
			MPS = 1
			MPPIO = 2 
			MPPI = 0
		elif Puo > 16 and Puo <= 24 and Pui > 36 and Pui <= 60:
			MPS = 1
			MPPIO = 2 
			MPPI = 1
		elif Puo > 16 and Puo <= 24 and Pui > 60 and Pui <= 84:
			MPS = 1
			MPPIO = 2 
			MPPI = 2
		elif Puo > 24 and Puo <= 32 and Pui > 0 and Pui <= 48:
			MPS = 1
			MPPIO = 3 
			MPPI = 0
		elif Puo > 24 and Puo <= 32 and Pui > 48 and Pui <= 72:
			MPS = 1
			MPPIO = 3 
			MPPI = 1
		elif Puo > 24 and Puo <= 32 and Pui > 72 and Pui <= 96:
			MPS = 1
			MPPIO = 3 
			MPPI = 2
		elif Puo > 32 and Puo <= 40 and Pui > 0 and Pui <= 60:
			MPS = 1
			MPPIO = 4 
			MPPI = 0
		elif Puo > 32 and Puo <= 40 and Pui > 60 and Pui <= 84:
			MPS = 1
			MPPIO = 4 
			MPPI = 1
		elif Puo > 32 and Puo <= 40 and Pui > 84 and Pui <= 108:
			MPS = 1
			MPPIO = 4 
			MPPI = 2
		elif Puo > 40 and Puo <= 48 and Pui > 0 and Pui <= 72:
			MPS = 1
			MPPIO = 5 
			MPPI = 0
		elif Puo > 40 and Puo <= 48 and Pui > 72 and Pui <= 96:
			MPS = 1
			MPPIO = 5 
			MPPI = 1
		elif Puo > 40 and Puo <= 48 and Pui > 96 and Pui <= 120:
			MPS = 1
			MPPIO = 5 
			MPPI = 2
		wsBG.cell(BG_max_row_1, 2).value = wskl.cell(i,2).value
		wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
		BG_max_row_1 = wsBG.max_row + 1
		if MPS >= 1:
			wordfinder('MP-S', MPS, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		if MPPIO >= 1:
			wordfinder('MPP-IO-U', MPPIO, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		if MPPI >= 1:
			wordfinder('MPP-I', MPPI, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		# Tìm giá tủ
		SL_Controler = MPS + MPPI + MPPIO
		if SL_Controler == 1:
			wordfinder('CP-01', 1, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		elif SL_Controler == 2:
			wordfinder('CP-02', 1, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		elif SL_Controler == 3:
			wordfinder('CP-03', 1, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		elif SL_Controler == 4:
			wordfinder('CP-04', 1, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		elif SL_Controler == 5:
			wordfinder('CP-05', 1, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		elif SL_Controler == 6:
			wordfinder('CP-06', 1, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		elif SL_Controler == 7:
			wordfinder('CP-07', 1, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
		elif SL_Controler == 8:
			wordfinder('CP-08', 1, BG_max_row_1)
			BG_max_row_1 = wsBG.max_row + 1
	#Cảm biến
	wsBG.cell(BG_max_row_1, 2).value = "Thiết bị trường"
	wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
	BG_max_row_1 = wsBG.max_row + 1

	# Hệ Chiller
	# SL cảm biến nhiệt độ nước
	SL_WTemp = wscb["D16"].value
	if SL_WTemp >= 1:
		wordfinder("TSAPA24", SL_WTemp, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1 
	# SL công tắc dòng chảy
	SL_WFS = wscb["D17"].value
	if SL_WFS >= 1:
		wordfinder("WFS", SL_WFS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL công tắc chênh áp suất nước
	SL_WDPS = wscb["D18"].value
	if SL_WDPS >= 1:
		wordfinder("WDPS", SL_WDPS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL cảm biến chênh áp suất nước
	SL_WDPT = wscb["D19"].value
	if SL_WDPT >= 1:
		wordfinder("WDPT", SL_WDPT, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL cảm biến lưu lượng nước
	SL_IEF = wscb["D20"].value
	if SL_IEF >= 1:
		wordfinder("FT", SL_IEF, BG_max_row_1)

	# Hệ AHU, PAU, FCU
	# SL cảm biến nhiệt độ ống gió
	SL_DTE = wscb["D24"].value
	if SL_DTE >= 1:
		wordfinder("DTE", SL_DTE, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL cảm biến nhiệt độ, độ ẩm lắp ống gió
	SL_DTH = wscb["D25"].value
	if SL_DTH >= 1:
		wordfinder("DTH", SL_DTH, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL cảm biến nhiệt độ lắp kho lạnh
	SL_TSFL = wscb["D26"].value
	if SL_TSFL >= 1:
		wordfinder("TSFL", SL_TSFL, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL công tắc gió
	SL_ADPS = wscb["D27"].value
	if SL_ADPS >= 1:
		wordfinder("ADPS", SL_ADPS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL công tắc mức nước
	SL_LS = wscb["D28"].value
	if SL_LS >= 1:
		wordfinder("LS", SL_LS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL công tắc báo lọc dơ
	SL_Fil = wscb["D29"].value
	if SL_Fil >= 1:
		wordfinder("ADPS", SL_Fil, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL công tắc quá nhiệt
	SL_OH = wscb["D30"].value
	if SL_OH >= 1:
		wordfinder("OH", SL_OH, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL bộ dẫn động muldating
	SL_MD_MO = wscb["D31"].value
	if SL_MD_MO >= 1:
		wordfinder("MD_MODUL", SL_MD_MO, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL bộ dẫn động On/Off
	SL_MD_OF = wscb["D32"].value
	if SL_MD_OF >= 1:
		wordfinder("MD_ON-OFF", SL_MD_OF, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

		BG_max_row_1 = wsBG.max_row + 1
	# SL Cảm biến chênh áp suất không khí
	SL_ADPT = wscb["D33"].value
	if SL_ADPT >= 1:
		wordfinder("ADPT", SL_ADPT, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

		BG_max_row_1 = wsBG.max_row + 1

	# SL Cảm biến chênh áp suất không khí
	SL_RFDPT = wscb["D34"].value
	if SL_RFDPT >= 1:
		wordfinder("ADPT", SL_RFDPT, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

		BG_max_row_1 = wsBG.max_row + 1

	# SL Cảm biến vận tốc gió
	SL_AV = wscb["D34"].value
	if SL_AV >= 1:
		wordfinder("AV", SL_AV, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
		BG_max_row_1 = wsBG.max_row + 1

	#EMS
	# SL cảm biến nhiệt độ, độ ẩm lắp phòng
	SL_RTH = wscb["D36"].value
	if SL_RTH >= 1:
		wordfinder("RTH", SL_RTH, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL cảm biến chênh áp phòng
	SL_DPT = wscb["D37"].value
	if SL_DPT >= 1:
		wordfinder("DPT", SL_DPT, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# SL cảm biến nhiệt độ phòng
	SL_RTE = wscb["D38"].value
	if SL_RTE >= 1:
		wordfinder("RTE", SL_RTE, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	#Trạm BMS
	wsBG.cell(BG_max_row_1, 2).value = "Trạm vận hành BMS"
	wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
	BG_max_row_1 = wsBG.max_row + 1

	# Bộ chia mạng 24 cổng
	SL_SW24BMS = wsBMS["D16"].value
	if SL_SW24BMS >= 1:
		wordfinder("SW-24", SL_SW24BMS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Bộ chia mạng 16 cổng
	SL_SW16BMS = wsBMS["D17"].value
	if SL_SW16BMS >= 1:
		wordfinder("SW-16", SL_SW16BMS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Máy tính vận hành BMS
	SL_WSCOM = wsBMS["D18"].value
	if SL_WSCOM >= 1:
		wordfinder("WS-COM", SL_WSCOM, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Máy in A4 BMS
	SL_BMSPRI = wsBMS["D19"].value
	if SL_BMSPRI >= 1:
		wordfinder("WS-PRINTER", SL_BMSPRI, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Bộ lưu điện UPS
	SL_BMSUPS = wsBMS["D20"].value
	if SL_BMSUPS >= 1:
		wordfinder("WS-UPS", SL_BMSUPS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	#Trạm EMS
	wsBG.cell(BG_max_row_1, 2).value = "Trạm vận hành EMS"
	wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
	BG_max_row_1 = wsBG.max_row + 1

	# Bộ chia mạng 24 cổng
	SL_SW24EMS = wsBMS["D24"].value
	if SL_SW24EMS >= 1:
		wordfinder("SW-24", SL_SW24EMS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Bộ chia mạng 16 cổng
	SL_SW16EMS = wsBMS["D25"].value
	if SL_SW16EMS >= 1:
		wordfinder("SW-16", SL_SW16EMS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Máy tính vận hành EMS
	SL_WSSERVER = wsBMS["D26"].value
	if SL_WSSERVER >= 1:
		wordfinder("WS-SERVER", SL_WSSERVER, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Máy in A4 EMS
	SL_EMSPRI = wsBMS["D27"].value
	if SL_EMSPRI >= 1:
		wordfinder("WS-PRINTER", SL_EMSPRI, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Bộ lưu điện UPS
	SL_EMSUPS = wsBMS["D28"].value
	if SL_EMSUPS >= 1:
		wordfinder("WS-UPS", SL_EMSUPS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	#Phần mềm điều khiển
	wsBG.cell(BG_max_row_1, 2).value = "Phần mềm điều khiển"
	wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
	BG_max_row_1 = wsBG.max_row + 1

	# Bộ điều khiển mạng
	SL_WEB = wsBMS["D32"].value
	if SL_WEB >= 1:
		wordfinder("MPW-C", SL_WEB, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Cảnh báo SMS
	SL_SMS = wsBMS["D33"].value
	if SL_SMS >= 1:
		wordfinder("SMS", SL_SMS, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Phần mềm CFR 21 Part 11
	SL_Software = wsBMS["D34"].value
	if SL_Software >= 1:
		wordfinder("RC-WB3", SL_Software, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# RC-Reporter
	SL_R3 = wsBMS["D35"].value
	if SL_R3 >= 1:
		wordfinder("RC-R3", SL_R3, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Archived
	SL_AR3 = wsBMS["D36"].value
	if SL_AR3 >= 1:
		wordfinder("RC-AR3", SL_AR3, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1
	# Tính tiền T&C
	Total_point = 0
	Total_graphics = 0
	for i in range(15,wskl.max_row+1):
		for j in range(3,7):
			Total_point = Total_point + wskl.cell(i,j).value
			Total_graphics = wskl.max_row+1 - 16
			Total_TC = Total_point + Total_graphics * 2
	#Chèn tiền T&C
	wsBG.cell(BG_max_row_1, 2).value = "Cài đặt, lập trình"
	wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
	BG_max_row_1 = wsBG.max_row + 1
	T_C = wsDM_T_C["C2"].value
	T_C1 = T_C * Total_TC

	def Laptrinh(searchString, Dongia):
		for i in range(1, wsDonGia.max_row + 1):
			for j in range(1, wsDonGia.max_column + 1):
				if searchString == wsDonGia.cell(i,j).value:
					wsDonGia.cell(i,j+4).value = Dongia
	Laptrinh('T_C',T_C1)
	#print(T_C1)

	if T_C1 >= 1:
		wordfinder('T_C', 1, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	wsBG.cell(BG_max_row_1, 2).value = "Sensor Calibration"
	wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
	BG_max_row_1 = wsBG.max_row + 1
	SL_CALIB = SL_RTH + SL_DPT
	wordfinder('CALIB', SL_CALIB, BG_max_row_1)
	BG_max_row_1 = wsBG.max_row + 1

	#Vật tư thi công
	wsBG.cell(BG_max_row_1, 2).value = "Vật tư thi công"
	wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
	BG_max_row_1 = wsBG.max_row + 1

	# Cáp mạng CAT6
	SL_CAT6 = wsVTTC["D16"].value
	if SL_CAT6 >= 1:
		wordfinder("CAT6", SL_CAT6, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	# Cáp AWG18 2 pair
	SL_AWG18_2 = wsVTTC["D17"].value
	if SL_AWG18_2 >= 1:
		wordfinder("AWG18_2", SL_AWG18_2, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	# Cáp AWG18 1 pair
	SL_AWG18_1 = wsVTTC["D18"].value
	if SL_AWG18_1 >= 1:
		wordfinder("AWG18_1", SL_AWG18_1, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	# Cáp Cu/PVC
	SL_Cu_PVC = wsVTTC["D19"].value
	if SL_Cu_PVC >= 1:
		wordfinder("Cu_PVC", SL_Cu_PVC, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	# Ống điện PVC D20
	SL_PVC_D20 = wsVTTC["D20"].value
	if SL_PVC_D20 >= 1:
		wordfinder("PVC_D20", SL_PVC_D20, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	# Ống điện GI D20
	SL_EMT_D20 = wsVTTC["D21"].value
	if SL_EMT_D20 >= 1:
		wordfinder("EMT_D20", SL_EMT_D20, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	# Ống điện GI D25
	SL_EMT_D25 = wsVTTC["D22"].value
	if SL_EMT_D25 >= 1:
		wordfinder("EMT_D25", SL_EMT_D25, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	# Trunking 200x100
	SL_TRUNKING200_100 = wsVTTC["D23"].value
	if SL_TRUNKING200_100 >= 1:
		wordfinder("TRUNKING200_100", SL_TRUNKING200_100, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	# Trunking 100x100
	SL_TRUNKING100_100 = wsVTTC["D24"].value
	if SL_TRUNKING100_100 >= 1:
		wordfinder("TRUNKING100_100", SL_TRUNKING100_100, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	# Vật tư phụ
	SL_VTP = wsVTTC["D25"].value
	if SL_VTP >= 1:
		wordfinder("VTP", SL_VTP, BG_max_row_1)
		BG_max_row_1 = wsBG.max_row + 1

	for codeVT in 'VT':
		sumVT=0
		for i in range(1,wsBG.max_row):
			if (wsBG.cell(i,12).value == codeVT):
				wsBG.cell(wsBG.max_row,14).value= ((wsBG.cell(i,14).value + wsBG.cell(i,15).value)* wsBG.cell(i,5).value)*0.4 + wsBG.cell(wsBG.max_row,14).value
				wsBG.cell(wsBG.max_row,15).value = 0.55 * wsBG.cell(wsBG.max_row,14).value + wsBG.cell(wsBG.max_row,15).value
	#print(Total_point)
	#print(Total_graphics)

	#Tính tổng giá tiền
	str71 = "=sum(J6:"
	str72 = str(wsBG.cell(BG_max_row_1-1, 10))
	str73 = ")"
	str74 = str71 + str72[12:-1] + str73
	wsBG.cell(BG_max_row_1, 10).value = str74
	wsBG.cell(BG_max_row_1, 10).font = Font(bold=True)
	wsBG.cell(BG_max_row_1, 10).number_format = '#,##0'
	# Merge ô tính tổng
	str80 = str(BG_max_row_1)
	str81 = "B"
	str82 = ":"
	str83 = "I"
	str84 = str81 + str80 + str82 + str83 + str80
	wsBG.merge_cells(str84)
	wsBG.cell(BG_max_row_1, 2).value = "TỔNG (CHƯA BAO GỒM VAT)"
	wsBG.cell(BG_max_row_1, 2).font = Font(bold=True)
	# Định dạng row
	for row in range(wsBG.max_row):
		wsBG.row_dimensions[row].height = 30
	# Kẻ border
	for row in range(6, wsBG.max_row+1):
		for col in range(1, wsBG.max_column+1):
			wsBG.cell(row, col).border = thin_border
	# wrap_text
	for row in range(6, wsBG.max_row):
		for col in range(4, 6):
			wsBG.cell(row, col).alignment = Alignment(horizontal="center")
	for row in range(6, wsBG.max_row):
		for col in range(2, 4):
			wsBG.cell(row, col).alignment = Alignment(wrap_text=True)
	#Định dạng số
	for row in range(6, wsBG.max_row):
		for col in range(6, wsBG.max_column-3):
			wsBG.cell(row, col).number_format = '#,##0'
	for row in range(6, wsBG.max_row):
		for col in range(wsBG.max_column-1, wsBG.max_column+1):
			wsBG.cell(row, col).number_format = '#,##0'

	#tong sum code vat tu

	wsSUM = wb_BG.create_sheet('sheet2')
	wsSUM.title = "SUM_CODE"

	wsSUM.append(['HẠNG MỤC', 'TỔNG TIỀN'])
	code_list = ['C', 'CP','CB','VT','WS', 'HC', 'TC']

	tongVT=0
	for code in code_list:
		sumTT=0
		for i in range(1,wsBG.max_row):
			if (wsBG.cell(i,12).value == code):
				sumTT= (wsBG.cell(i,14).value + wsBG.cell(i,15).value)* wsBG.cell(i,5).value + sumTT				
		wsSUM.append([code, sumTT])
		if(code== 'VT'):
			tongVT = sumTT

	#ghi ten du an 
	duan='Dự án : '
	DA_name = duan + DA_name
	wsBG.cell(2,1).value = DA_name
	wsBG.cell(wsBG.max_row-1,14).value = tongVT
	wsBG.cell(wsBG.max_row-1,15).value = tongVT

	#save

	wb_BG.save(save_path+"\Bao_gia_BMS.xlsx")
	print("THE QUOTATION IS COMPLETED")
	time.sleep(1)

